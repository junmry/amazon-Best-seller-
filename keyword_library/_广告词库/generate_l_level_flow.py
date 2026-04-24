#!/usr/bin/env python3
"""
L级别投流词分层脚本
按 Recliner ABA五层流量模型 对全部搜索词库进行投流词分层

处理8个类目：吧凳、户外家具、桌布、火坑、玄关桌（有排名）、沙发、椅子、人体工学椅（无排名）
输出到 keyword_library/_广告词库/L级别投流词/
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json

# ============================================================
# 分层标准定义
# ============================================================

LAYERS = [
    {"name": "L1锚定层", "rank_min": 1, "rank_max": 1000, "vol_min": 5000, "vol_max": float('inf'),
     "flow_attr": "高流量核心入口", "strategy": "精准匹配+高竞价抢Top3", "budget_pct": "40%", "match_type": "精准匹配", "bid": "高竞价"},
    {"name": "L2主战场", "rank_min": 1001, "rank_max": 30000, "vol_min": 500, "vol_max": 4999,
     "flow_attr": "高流量主战场", "strategy": "词组匹配，分时段加价放量", "budget_pct": "35%", "match_type": "词组匹配", "bid": "中竞价"},
    {"name": "L3验证区", "rank_min": 30001, "rank_max": 150000, "vol_min": 100, "vol_max": 499,
     "flow_attr": "中流量验证区", "strategy": "广泛匹配低竞价，跑转化养数据", "budget_pct": "20%", "match_type": "广泛匹配", "bid": "低竞价"},
    {"name": "L4长尾区", "rank_min": 150001, "rank_max": 500000, "vol_min": 10, "vol_max": 99,
     "flow_attr": "低流量长尾", "strategy": "广泛匹配捡漏，找稳定出单词", "budget_pct": "4%", "match_type": "广泛匹配", "bid": "低竞价"},
    {"name": "L5探索区", "rank_min": 500001, "rank_max": float('inf'), "vol_min": 0, "vol_max": 9,
     "flow_attr": "极低流量探索", "strategy": "自动广告跑词，定期否定/打捞", "budget_pct": "1%", "match_type": "自动广告", "bid": "最低竞价"},
]

OUTPUT_DIR = "/root/.openclaw/workspace/keyword_library/_广告词库/L级别投流词"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# 样式定义
# ============================================================

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LAYER_FILLS = {
    "L1锚定层": PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid"),
    "L2主战场": PatternFill(start_color="F79646", end_color="F79646", fill_type="solid"),
    "L3验证区": PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid"),
    "L4长尾区": PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid"),
    "L5探索区": PatternFill(start_color="8064A2", end_color="8064A2", fill_type="solid"),
}
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ============================================================
# 读取各品类源数据
# ============================================================

def read_excel_data(path, sheet_name):
    """读取Excel数据，返回 (header, rows)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return rows[0], rows[1:]

def classify_by_rank(rank):
    """按排名分层"""
    if rank is None or rank == '' or rank == 0:
        return None
    try:
        r = int(rank)
    except:
        return None
    if r <= 1000:
        return "L1锚定层"
    elif r <= 30000:
        return "L2主战场"
    elif r <= 150000:
        return "L3验证区"
    elif r <= 500000:
        return "L4长尾区"
    else:
        return "L5探索区"

def classify_by_volume(vol):
    """按搜索量分层（无排名数据时使用）"""
    if vol is None or vol == '':
        vol = 0
    try:
        v = int(vol)
    except:
        v = 0
    if v >= 5000:
        return "L1锚定层"
    elif v >= 500:
        return "L2主战场"
    elif v >= 100:
        return "L3验证区"
    elif v >= 10:
        return "L4长尾区"
    else:
        return "L5探索区"

def get_volume(row_dict, vol_keys):
    """从行字典中提取搜索量值"""
    for k in vol_keys:
        if k in row_dict and row_dict[k] not in (None, ''):
            try:
                return int(row_dict[k])
            except:
                pass
    return 0

def get_rank(row_dict):
    """从行字典中提取排名值"""
    if '排名' in row_dict and row_dict['排名'] not in (None, ''):
        try:
            return int(row_dict['排名'])
        except:
            pass
    return None

def make_row_dict(header, row):
    """将行数据转为字典"""
    return {h: v for h, v in zip(header, row) if h is not None}

def process_category(name, path, sheet, has_rank=True, vol_keys=None):
    """处理单个品类，生成分层Excel"""
    print(f"\n>>> 处理: {name}")
    header, rows = read_excel_data(path, sheet)
    if not header:
        print(f"  跳过: 无数据")
        return
    
    # 收集分层数据
    layer_data = {layer["name"]: [] for layer in LAYERS}
    
    for row in rows:
        rd = make_row_dict(header, row)
        
        # 获取关键词
        keyword = rd.get('关键词', '') or rd.get('Keyword', '') or rd.get('keyword', '')
        if not keyword or keyword == '关键词':
            continue
        
        # 获取搜索量
        vol = get_volume(rd, vol_keys or ['周搜索量', '搜索量', '月搜索量'])
        
        # 获取排名
        rank = get_rank(rd) if has_rank else None
        
        # 分层
        if has_rank and rank is not None:
            layer_name = classify_by_rank(rank)
        else:
            layer_name = classify_by_volume(vol)
        
        if layer_name is None:
            continue
        
        layer_data[layer_name].append({
            '关键词': keyword,
            '翻译': rd.get('翻译', ''),
            '搜索量': vol,
            '排名': rank if rank is not None else 'N/A',
            '相关性': rd.get('相关性-自动', '') or rd.get('相关性', ''),
        })
    
    # 创建输出Excel
    out_path = os.path.join(OUTPUT_DIR, f"{name}_L级别投流词分层.xlsx")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # 删除默认sheet
    
    # ===== Sheet 0: 分层概览 =====
    ws_overview = wb_out.create_sheet("0-分层概览")
    overview_headers = ["ABA分层", "排名区间", "关键词数量", "总搜索量", "流量属性", "投放策略", "预算占比建议", "匹配方式", "竞价建议"]
    for col, h in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BORDER
    
    for i, layer in enumerate(LAYERS, 2):
        data = layer_data[layer["name"]]
        total_vol = sum(d['搜索量'] for d in data)
        rank_range = f"{layer['rank_min']:,}-{layer['rank_max']:,}" if layer['rank_max'] != float('inf') else f"{layer['rank_min']:,}+"
        
        row_data = [
            layer["name"],
            rank_range,
            len(data),
            total_vol,
            layer["flow_attr"],
            layer["strategy"],
            layer["budget_pct"],
            layer["match_type"],
            layer["bid"],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_overview.cell(row=i, column=col, value=val)
            cell.fill = LAYER_FILLS[layer["name"]]
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER
    
    # 调整列宽
    for col in range(1, len(overview_headers) + 1):
        ws_overview.column_dimensions[get_column_letter(col)].width = 18
    ws_overview.column_dimensions['F'].width = 35
    
    # ===== Sheets 1-5: 各层级详情 =====
    detail_headers = ["关键词", "翻译", "搜索量", "排名", "相关性", "投放状态", "投放策略建议"]
    
    for layer in LAYERS:
        data = layer_data[layer["name"]]
        # 按搜索量降序排列
        data.sort(key=lambda x: x['搜索量'], reverse=True)
        
        ws = wb_out.create_sheet(f"{layer['name'][0:2]}-{layer['name']}")
        
        # 标题行
        for col, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER
        
        # 数据行
        for ridx, d in enumerate(data, 2):
            row_vals = [
                d['关键词'],
                d['翻译'],
                d['搜索量'],
                d['排名'],
                d['相关性'],
                "待投放",
                layer['strategy'],
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=ridx, column=col, value=val)
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN if col in (3, 4) else Alignment(horizontal="left", vertical="center")
        
        # 列宽
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 35
        
        # 冻结首行
        ws.freeze_panes = "A2"
    
    wb_out.save(out_path)
    print(f"  完成: {out_path}")
    print(f"  总计词数: {sum(len(v) for v in layer_data.values())}")
    for layer in LAYERS:
        print(f"    {layer['name']}: {len(layer_data[layer['name']])}个词")


# ============================================================
# 主执行
# ============================================================

if __name__ == "__main__":
    # 1. 有排名的品类
    process_category(
        "吧凳_Barstool",
        "/root/.openclaw/workspace/keyword_library/吧凳_Barstool/02_拆词结果/吧凳关键词拆词_完整层级V2.xlsx",
        "源文件",
        has_rank=True,
        vol_keys=['周搜索量'],
    )
    
    process_category(
        "户外家具_Outdoor",
        "/root/.openclaw/workspace/keyword_library/户外家具_Outdoor/02_拆词结果/户外家具关键词拆词_完整层级V2.xlsx",
        "源文件",
        has_rank=True,
        vol_keys=['搜索量'],
    )
    
    process_category(
        "桌布_Tablecloth",
        "/root/.openclaw/workspace/keyword_library/桌布_Tablecloth/02_拆词结果/桌布关键词拆词_完整层级V2.xlsx",
        "源文件",
        has_rank=True,
        vol_keys=['周搜索量'],
    )
    
    process_category(
        "火坑_FirePit",
        "/root/.openclaw/workspace/keyword_library/火坑_FirePit/02_拆词结果/firepit_keywords_split.xlsx",
        "筛选",
        has_rank=True,
        vol_keys=['月搜索量'],
    )
    
    process_category(
        "玄关桌_ConsoleTable",
        "/root/.openclaw/workspace/keyword_library/玄关桌_ConsoleTable/02_拆词结果/SofaConsoleTable_关键词拆词.xlsx",
        "源文件",
        has_rank=True,
        vol_keys=['搜索量'],
    )
    
    # 2. 无排名的品类（按搜索量映射）
    process_category(
        "沙发_Sofa",
        "/root/.openclaw/workspace/keyword_library/沙发_Sofa/02_拆词结果/沙发关键词拆词_完整层级V2.xlsx",
        "源文件",
        has_rank=False,
        vol_keys=['搜索量'],
    )
    
    process_category(
        "椅子_Chair",
        "/root/.openclaw/workspace/keyword_library/椅子_Chair/02_拆词结果/椅子关键词拆词_完整层级.xlsx",
        "源文件",
        has_rank=False,
        vol_keys=['搜索量'],
    )
    
    process_category(
        "人体工学椅_Ergonomic",
        "/root/.openclaw/workspace/keyword_library/人体工学椅_Ergonomic/02_拆词结果/人体工学办公椅关键词拆词.xlsx",
        "源文件",
        has_rank=False,
        vol_keys=['搜索量'],
    )
    
    print(f"\n{'='*60}")
    print("全部处理完成！")
    print(f"输出目录: {OUTPUT_DIR}")

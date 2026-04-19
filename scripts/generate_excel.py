#!/usr/bin/env python3
"""
生成金融从业资格考试Excel规划表
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 创建工作簿
wb = openpyxl.Workbook()

# 样式定义 - 使用正确的PatternFill语法
header_fill = PatternFill(patternType="solid", fgColor="366092")
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
title_font = Font(name="微软雅黑", size=14, bold=True)
normal_font = Font(name="微软雅黑", size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_cell_style(cell, font=normal_font, fill=None, align="left"):
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = border

# ==================== Sheet 1: 考试时间表 ====================
ws1 = wb.active
ws1.title = "考试时间表"

# 标题
ws1.merge_cells('A1:H1')
ws1['A1'] = "2026年金融从业资格考试时间表"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

# 表头
headers = ["考试名称", "考试类型", "考试时间", "报名时间", "报名网站", "科目", "费用", "成绩有效期"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, "center")

# 数据
data = [
    ["期货从业", "统考", "5月16日", "4月中下旬", "www.cfachina.org", "基础知识+法律法规", "65元/科", "长期"],
    ["期货从业", "专场", "9月19日", "待定", "www.cfachina.org", "三科可选", "65元/科", "长期"],
    ["期货从业", "专场", "11月21日", "待定", "www.cfachina.org", "三科可选", "65元/科", "长期"],
    ["基金从业", "统考", "5月23日", "4月中下旬", "www.amac.org.cn", "科1必考+科2/3选考", "61元/科", "4年"],
    ["基金从业", "统考", "11月28日", "10月中下旬", "www.amac.org.cn", "科1必考+科2/3选考", "61元/科", "4年"],
    ["证券从业", "统考", "6月27日", "5月中下旬", "www.sac.net.cn", "金融市场基础+法律法规", "61元/科", "36个月"],
    ["证券从业", "统考", "9月19日", "8月中下旬", "www.sac.net.cn", "金融市场基础+法律法规", "61元/科", "36个月"],
    ["证券从业", "专场", "4月18日", "3月中下旬", "www.sac.net.cn", "两科", "61元/科", "36个月"],
    ["证券从业", "专场", "11月", "10月中下旬", "www.sac.net.cn", "两科", "61元/科", "36个月"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, align="center" if col_idx <= 5 else "left")

# 设置列宽
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 22
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 12

# ==================== Sheet 2: 备考时间线 ====================
ws2 = wb.create_sheet("备考时间线")

ws2.merge_cells('A1:F1')
ws2['A1'] = "推荐备考时间线（稳妥版）"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

timeline_headers = ["阶段", "时间", "目标考试", "备考内容", "建议时长", "备注"]
for col, header in enumerate(timeline_headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, "center")

timeline_data = [
    ["准备期", "3月", "了解考情", "查看大纲，准备资料", "1周", "确定报考科目"],
    ["第一阶段", "4月", "期货从业", "基础知识+法规", "4-5周", "重点：套期保值、套利"],
    ["第二阶段", "5-6月", "证券从业", "金融基础+法规", "5-6周", "重点：股票、债券、衍生品"],
    ["第三阶段", "9-10月", "基金从业", "科1+科2", "6-8周", "重点：投资组合、估值"],
    ["冲刺", "考前2周", "全科", "刷真题+错题", "2周", "机位紧张，早报早安心"],
]

for row_idx, row_data in enumerate(timeline_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, align="center" if col_idx <= 3 else "left")

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 20

# ==================== Sheet 3: 各科重点分值 ====================
ws3 = wb.create_sheet("各科重点")

ws3.merge_cells('A1:E1')
ws3['A1'] = "各科目重点章节与分值分布"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")

# 证券从业
ws3['A3'] = "【证券从业】金融市场基础知识"
ws3['A3'].font = Font(name="微软雅黑", size=11, bold=True, color="366092")

sec_headers = ["章节", "内容", "分值", "难度", "重点标记"]
for col, header in enumerate(sec_headers, 1):
    cell = ws3.cell(row=4, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, "center")

sec_data = [
    ["第四章", "股票", "20分", "★★★", "股票估值、发行交易"],
    ["第七章", "金融衍生工具", "15分", "★★★", "期货、期权、互换"],
    ["第二章", "中国金融体系", "15分", "★★", "多层次资本市场"],
    ["第五章", "债券", "15分", "★★", "债券估值、风险"],
    ["第一章", "金融市场体系", "10分", "★", "基本概念"],
    ["第三章", "证券市场主体", "10分", "★", "中介机构"],
    ["第六章", "证券投资基金", "10分", "★", "基金类型"],
    ["第八章", "金融风险管理", "5分", "★", "风险类型"],
]

for row_idx, row_data in enumerate(sec_data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, align="center" if col_idx != 5 else "left")

# 基金从业
row_offset = 14
ws3.cell(row=row_offset, column=1, value="【基金从业】科目二：证券投资基金基础知识")
ws3.cell(row=row_offset, column=1).font = Font(name="微软雅黑", size=11, bold=True, color="366092")

for col, header in enumerate(sec_headers, 1):
    cell = ws3.cell(row=row_offset+1, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, "center")

fund_data = [
    ["第三章", "固定收益投资", "15分", "★★★", "久期、凸性、YTM"],
    ["第七章", "投资组合管理", "15分", "★★★", "CAPM、有效市场"],
    ["第二章", "权益投资", "15分", "★★", "股票估值方法"],
    ["第四章", "衍生工具", "10分", "★★", "远期、期货、期权"],
    ["第十章", "基金业绩评价", "10分", "★★", "夏普比率、阿尔法"],
    ["第一章", "投资管理基础", "10分", "★", "财务报表分析"],
    ["第九章", "投资风险管理", "10分", "★", "VaR、风险类型"],
]

for row_idx, row_data in enumerate(fund_data, row_offset+2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, align="center" if col_idx != 5 else "left")

# 期货从业
row_offset = 23
ws3.cell(row=row_offset, column=1, value="【期货从业】期货基础知识")
ws3.cell(row=row_offset, column=1).font = Font(name="微软雅黑", size=11, bold=True, color="366092")

for col, header in enumerate(sec_headers, 1):
    cell = ws3.cell(row=row_offset+1, column=col, value=header)
    set_cell_style(cell, header_font, header_fill, "center")

futures_data = [
    ["第三章", "套期保值", "20分", "★★★", "基差、套保效果计算"],
    ["第四章", "投机与套利", "20分", "★★★", "价差套利盈亏计算"],
    ["第二章", "期货合约与制度", "15分", "★★", "保证金、逐日盯市"],
    ["第六章", "金融期货", "15分", "★★", "国债期货、股指期货"],
    ["第五章", "期货期权", "10分", "★★", "希腊字母、盈亏平衡点"],
    ["第一章", "期货市场概述", "10分", "★", "发展历程"],
    ["第七章", "监管与风控", "10分", "★", "风险控制体系"],
]

for row_idx, row_data in enumerate(futures_data, row_offset+2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        set_cell_style(cell, align="center" if col_idx != 5 else "left")

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 30

# ==================== Sheet 4: 报名检查清单 ====================
ws4 = wb.create_sheet("报名清单")

ws4.merge_cells('A1:D1')
ws4['A1'] = "报名前检查清单"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal="center", vertical="center")

checklist = [
    ("基本条件", [
        "年满18周岁",
        "具有完全民事行为能力",
        "高中/大专及以上学历（证券需大专或高中+36个月经验）",
    ]),
    ("准备材料", [
        "身份证原件（有效期内）",
        "学历证明（毕业证编号或扫描件）",
        "近期白底证件照（JPG格式，30KB-100KB）",
        "电子邮箱（接收准考证和成绩）",
        "手机号码（接收验证码）",
    ]),
    ("报名流程", [
        "1. 登录官网注册账号",
        "2. 填写个人信息（姓名/身份证号不可更改，务必核对）",
        "3. 选择考试科目",
        "4. 选择考区城市",
        "5. 在线支付报名费",
        "6. 确认报名成功",
    ]),
    ("考前准备", [
        "考前一周打印准考证",
        "确认考试地点和时间",
        "准备身份证原件+准考证",
        "熟悉考场路线",
        "复习重点章节",
    ]),
    ("注意事项", [
        "⚠️ 机位有限，先报先得！",
        "⚠️ 缴费成功才算报名完成",
        "⚠️ 姓名和身份证号注册后不可修改",
        "⚠️ 连续两次缺考将被限制报考一次",
        "⚠️ 成绩有效期内申请从业资格",
    ]),
]

row = 3
for category, items in checklist:
    ws4.cell(row=row, column=1, value=category).font = Font(name="微软雅黑", size=11, bold=True, color="366092")
    ws4.merge_cells(f'A{row}:D{row}')
    row += 1
    for item in items:
        ws4.cell(row=row, column=1, value="☐")
        ws4.cell(row=row, column=2, value=item)
        ws4.merge_cells(f'B{row}:D{row}')
        row += 1
    row += 1

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 60

# 保存文件
output_path = "/root/.openclaw/workspace/金融从业资格考试规划表.xlsx"
wb.save(output_path)
print(f"✅ Excel文件已生成: {output_path}")

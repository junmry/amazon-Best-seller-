#!/usr/bin/env python3
import re
from collections import defaultdict
from openpyxl import load_workbook
import xlsxwriter

input_path = '/root/.openclaw/workspace/.kimi/downloads/19d8551c-f392-83d8-8000-0000c70c6520_variantExtendKeyword_lmifwzoqgiw1776058193762_1776058200940.xlsx'
output_path = '/root/.openclaw/workspace/吧凳关键词拆词_完整层级V2.xlsx'

# 读取输入
wb_in = load_workbook(input_path)
ws_in = wb_in.active

headers = [cell.value for cell in next(ws_in.iter_rows(min_row=2, max_row=2, max_col=ws_in.max_column))]
rows = []
keywords = []
volumes = []
for row in ws_in.iter_rows(min_row=3, values_only=True):
    row_data = list(row)
    rows.append(row_data)
    kw = row_data[1] if len(row_data) > 1 else None
    vol = row_data[10] if len(row_data) > 10 else 0
    if kw:
        keywords.append(str(kw).strip())
        volumes.append(int(vol) if vol else 0)

kw_data = [(k.lower(), k, v) for k, v in zip(keywords, volumes)]

dimensions = {
    '颜色': {
        '主词': ['white','black','brown','gray','grey','blue','red','green','beige','tan','cream','ivory','pink','purple','yellow','orange','navy','teal','turquoise','burgundy','maroon','charcoal','gold','silver','rose gold','coral','mint','lilac','lavender','emerald','sapphire','ruby','bronze','copper','champagne','blush','slate','rust','terracotta','sage','mustard','mauve','plum','peach','aqua','cobalt','crimson','scarlet','forest green','olive','khaki','chocolate','espresso','sand','stone','midnight blue','baby blue','sky blue','royal blue','lemon','tangerine','pumpkin','salmon','fuchsia','magenta','violet','indigo','rainbow','multi color','multicolor','ombre','gradient','tie dye','marble','clear','pastel','light','dark','natural','walnut','oak','cherry','mahogany','maple'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '尺码': {
        '主词': ['counter height','bar height','24 inch','26 inch','28 inch','30 inch','32 inch','34 inch','36 inch','24"','26"','28"','30"','32"','34"','36"','short','tall','extra tall','adjustable','low back','high back','full back','wide','narrow','small','large','big','compact','space saving','2 pack','4 pack','6 pack','set of 2','set of 4','set of 6'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '人群': {
        '主词': ['kids','kid','children','child','baby','toddler','family','adult','boys','boy','girls','girl','men','man','women','woman','teen','senior','short person','tall people','heavy duty','big and tall','overweight','plus size','petite'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '款式风格': {
        '主词': ['modern','contemporary','mid century','mid-century','farmhouse','rustic','vintage','industrial','antique','traditional','classic','elegant','luxury','minimalist','scandinavian','bohemian','coastal','french country','shabby chic','glamorous','art deco','retro','western','cowboy','tufted','wingback','barrel','saddle','bucket','cross back','x back','ladder back','slat back','open back','curved','square','round','upholstered','padded','cushioned','armless','with arms','with back','without back','backless','swivel','stationary','non swivel','non-swivel','fixed','stackable','nesting','folding','foldable','portable','outdoor','indoor','commercial','residential'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '材质': {
        '主词': ['wood','wooden','metal','steel','iron','wrought iron','aluminum','chrome','stainless steel','brass','bronze','copper','rattan','wicker','bamboo','seagrass','rope','cord','velvet','linen','fabric','leather','faux leather','pu leather','bonded leather','genuine leather','real leather','microfiber','suede','chenille','cotton','canvas','acrylic','plastic','resin','teak','acacia','pine','oak','walnut','rubberwood','mdf','particle board','laminate','glass','polycarbonate'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '场景': {
        '主词': ['kitchen','kitchen island','counter','breakfast bar','bar','home bar','patio','outdoor','deck','backyard','garden','balcony','porch','dining room','living room','basement','man cave','game room','pub','restaurant','cafe','bistro','office','workspace','commercial','indoor','rv','camper','apartment','small space'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '功能': {
        '主词': ['adjustable height','adjustable','swivel','360 swivel','rotating','footrest','foot rest','back support','ergonomic','lumbar support','cushioned','padded','upholstered','stackable','foldable','folding','nesting','storage','with wheels','wheeled','casters','easy assembly','quick assemble','no assembly','pre assembled','heavy duty','sturdy','durable','weight capacity','300 lb','400 lb','500 lb','weather resistant','waterproof','uv resistant','rust resistant','scratch resistant','non slip','anti slip','floor protector'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '价格': {
        '主词': ['cheap','affordable','budget','discount','inexpensive','low price','luxury','premium','expensive','high end','clearance','sale','best seller','top rated','value','deal','bargain','wholesale','bulk'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    },
    '品牌': {
        '主词': ['amazon','walmart','target','wayfair','ikea','home depot','lowe','costco','ashley','west elm','cb2','pottery barn','crate and barrel'],
        '副词': ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']
    }
}

core_products = ['bar stool','bar stools','stool','stools','counter stool','counter stools','barstool','barstools']

def extract_dimension_data(kw_data, dim_config):
    groups = defaultdict(list)
    main_words = sorted(dim_config['主词'], key=lambda x: len(x), reverse=True)
    compiled = {mw: re.compile(r'(?:^| )' + re.escape(mw.lower()) + r'(?:$| )') for mw in main_words}

    for main_word in main_words:
        pat = compiled[main_word]
        for kw_lower, kw_orig, vol in kw_data:
            if pat.search(kw_lower):
                副词 = ''
                次副词 = '无属性'
                for sub_word in dim_config['副词']:
                    if sub_word in kw_lower:
                        副词 = sub_word
                        break
                if not 副词:
                    for cp in core_products:
                        if cp in kw_lower:
                            副词 = cp
                            break
                if not 副词:
                    副词 = 'stool'
                groups[main_word].append({'副词': 副词, '次副词': 次副词, '搜索词': kw_orig, '搜索量': vol})

    for main_word in groups:
        groups[main_word].sort(key=lambda x: -x['搜索量'])

    results = []
    idx = 1
    for main_word in sorted(groups.keys()):
        items = groups[main_word]
        for i, item in enumerate(items):
            results.append([idx, main_word if i == 0 else '', item['副词'] if i == 0 else '', item['次副词'] if i == 0 else '', item['搜索词'], item['搜索量']])
            idx += 1
    return results

workbook = xlsxwriter.Workbook(output_path)
header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF', 'align': 'center', 'valign': 'vcenter'})

# 1. 源文件
ws_source = workbook.add_worksheet('源文件')
ws_source.write_row(0, 0, headers, header_fmt)
for i, row in enumerate(rows, 1):
    ws_source.write_row(i, 0, row)

# 2. 筛选
ws_filter = workbook.add_worksheet('筛选')
ws_filter.write_row(0, 0, ['序号'] + headers, header_fmt)
indexed_rows = [(i+1, r, int(r[10]) if len(r) > 10 and r[10] is not None else 0) for i, r in enumerate(rows)]
indexed_rows.sort(key=lambda x: -x[2])
for i, (idx, row, vol) in enumerate(indexed_rows, 1):
    ws_filter.write_row(i, 0, [idx] + row)

# 3. 筛选后词
all_words = defaultdict(lambda: {'count': 0, 'volume': 0})
for kw, vol in zip(keywords, volumes):
    words = re.findall(r"[a-z']+", kw.lower())
    for w in words:
        if len(w) > 1:
            all_words[w]['count'] += 1
            all_words[w]['volume'] += vol

ws_words = workbook.add_worksheet('筛选后词')
ws_words.write_row(0, 0, ['词', '出现次数', '搜索量'], header_fmt)
for i, (w, stats) in enumerate(sorted(all_words.items(), key=lambda x: -x[1]['count']), 1):
    ws_words.write_row(i, 0, [w, stats['count'], stats['volume']])

# 4-12. 各维度
dim_stats = {}
for dim_name, dim_config in dimensions.items():
    data = extract_dimension_data(kw_data, dim_config)
    dim_stats[dim_name] = len(data)
    if data:
        sheet_name = dim_name[:31]
        ws_dim = workbook.add_worksheet(sheet_name)
        ws_dim.write_row(0, 0, ['序号', '主词', '副词', '次副词', '搜索词', '搜索量'], header_fmt)
        for i, row in enumerate(data, 1):
            ws_dim.write_row(i, 0, row)

workbook.close()

print('吧凳关键词拆词 V2 完成！')
print(f'文件保存至: {output_path}')
print(f'总关键词数: {len(keywords)}')
print(f'总周搜索量: {sum(volumes):,}')
print('各维度统计:')
for dim_name, count in dim_stats.items():
    print(f'  {dim_name}: {count} 条')

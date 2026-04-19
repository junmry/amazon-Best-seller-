#!/usr/bin/env python3
import re
from collections import defaultdict
from openpyxl import load_workbook
import xlsxwriter

input_path = '/root/.openclaw/workspace/.kimi/downloads/19d84f82-5e42-8e00-8000-0000e41198bc_variantExtendKeyword_lhcvmnlozxg1776050980964_1776050994969.xlsx'
output_path = '/root/.openclaw/workspace/桌布关键词拆词_完整层级V2.xlsx'

# 读取输入（不用 read_only，以确保 max_row 准确；直接用 iter_rows 加速）
wb_in = load_workbook(input_path)
ws_in = wb_in.active

# 表头在 Row 2
raw_headers = [cell.value for cell in next(ws_in.iter_rows(min_row=2, max_row=2, max_col=ws_in.max_column))]
# 截断末尾 None
headers = [h for h in raw_headers if h is not None]
while len(raw_headers) > len(headers) and raw_headers[len(headers)] is None:
    pass
headers = raw_headers

rows = []
keywords = []
volumes = []
for r_idx, row in enumerate(ws_in.iter_rows(min_row=3, values_only=True), 3):
    row_data = list(row)
    rows.append(row_data)
    kw = row_data[1] if len(row_data) > 1 else None
    vol = row_data[10] if len(row_data) > 10 else 0
    if kw:
        keywords.append(str(kw).strip())
        volumes.append(int(vol) if vol else 0)

kw_data = [(k.lower(), k, v) for k, v in zip(keywords, volumes)]

dimensions = {
    '颜色': {
        '主词': ['white','black','brown','gray','grey','blue','red','green','beige','tan','cream','ivory','pink','purple','yellow','orange','navy','teal','turquoise','burgundy','maroon','charcoal','gold','silver','rose gold','coral','mint','lilac','lavender','emerald','sapphire','ruby','bronze','copper','champagne','blush','slate','rust','terracotta','sage','mustard','mauve','plum','peach','aqua','cobalt','crimson','scarlet','forest green','olive','khaki','chocolate','espresso','sand','stone','midnight blue','baby blue','sky blue','royal blue','lemon','tangerine','pumpkin','salmon','fuchsia','magenta','violet','indigo','rainbow','multi color','multicolor','ombre','gradient','tie dye','marble','clear','pastel','light'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths','cover','runner','napkins','placemat','placemats']
    },
    '尺码': {
        '主词': ['rectangle','rectangular','round','oval','square','oblong','6ft','6 ft','8ft','8 ft','60 inch','70 inch','90 inch','120 inch','132 inch','6 foot','8 foot','60x102','60x84','52x70','70x120','90x132','60 round','120 round','small','large','big','long','extra long','oversized','standard','fitted','elastic','stretch','ft','foot'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths']
    },
    '人群': {
        '主词': ['kids','kid','children','child','baby','toddler','family','adult','boys','boy','girls','girl','men','man','women','woman','teen','senior'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','party','shower','birthday']
    },
    '款式风格节日': {
        '主词': ['rustic','vintage','modern','floral','striped','plaid','checkered','gingham','buffalo check','geometric','plain','solid','printed','patterned','embroidered','damask','bohemian','farmhouse','elegant','luxury','simple','classic','contemporary','minimalist','moroccan','tropical','nautical','coastal','shabby chic','french country','industrial','art deco','mid century','victorian','romantic','glamorous','casual','formal','festive','seasonal','christmas','holiday','halloween','thanksgiving','easter','valentine','wedding','birthday','party','banquet','event','baby shower','bridal shower','anniversary','graduation','quinceanera','baptism','communion','fiesta','patriotic','fall','autumn','winter','spring','summer','harvest','pumpkin','snowflake','reindeer','santa','gingerbread','candy cane','poinsettia','themed','theme','princess','minnie','mickey','disney','mouse','winnie','pooh','demon','goose','bunny','kpop','car','gender reveal','reveal'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths','decorations','decor','party','supplies']
    },
    '材质': {
        '主词': ['polyester','cotton','linen','lace','vinyl','plastic','fabric','burlap','velvet','satin','spandex','stretch','sequin','taffeta','organza','faux leather','leather','denim','canvas','mesh','tulle','silk','rayon','chiffon','jacquard','damask','brocade','knit','woven','non woven','paper','peva','pvc','flannel','fleece','suede','wool','jute','hemp','organic','natural','synthetic','disposable','reusable'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths']
    },
    '场景': {
        '主词': ['dining room','kitchen','picnic','outdoor','indoor','restaurant','hotel','party','wedding','banquet','event','buffet','coffee table','patio','camping','bbq','conference','trade show','birthday','baby shower','bridal shower','office','home','commercial','everyday','special occasion','formal dinner','casual dining','outdoor dining','outdoor party','backyard','garden','poolside','beach','rv','cafeteria','bistro','catering','desk','bar','yard'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths','decorations','decor']
    },
    '功能': {
        '主词': ['waterproof','water resistant','spillproof','spill proof','stain resistant','wrinkle free','wrinkle resistant','washable','reusable','disposable','elastic','fitted','stretchable','non slip','heat resistant','oil proof','scratch resistant','easy clean','wipeable','durable','heavy duty','antibacterial','antimicrobial','fade resistant','uv resistant','tear resistant','flame retardant','fire resistant','insulated','padded','quilted','protective','decorative','protector','clear','folding','foldable'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths']
    },
    '价格': {
        '主词': ['cheap','affordable','budget','discount','luxury','premium','expensive','inexpensive','clearance','sale','wholesale','bulk'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers','party','decorations']
    },
    '品牌': {
        '主词': ['amazon','walmart','target','wayfair','ikea'],
        '副词': ['tablecloth','table cloth','tablecloths','table cover','table covers']
    }
}

core_products = ['tablecloth','table cloth','tablecloths','table cover','table covers','cloth','cloths','cover','runner','napkins','placemat','placemats']

def extract_dimension_data(kw_data, dim_config):
    groups = defaultdict(list)
    main_words = sorted(dim_config['主词'], key=lambda x: len(x), reverse=True)
    compiled = {mw: re.compile(r'(?:^| )' + re.escape(mw.lower()) + r'(?:$| )') for mw in main_words}

    for main_word in main_words:
        pat = compiled[main_word]
        for kw_lower, kw_orig, vol in kw_data:
            if pat.search(kw_lower):
                副词 = ''
                次副词 = '无属性'
                for sub_word in dim_config['副词']:
                    if sub_word in kw_lower:
                        副词 = sub_word
                        break
                if not 副词:
                    for cp in core_products:
                        if cp in kw_lower:
                            副词 = cp
                            break
                if not 副词:
                    副词 = 'tablecloth'
                groups[main_word].append({'副词': 副词, '次副词': 次副词, '搜索词': kw_orig, '搜索量': vol})

    for main_word in groups:
        groups[main_word].sort(key=lambda x: -x['搜索量'])

    results = []
    idx = 1
    for main_word in sorted(groups.keys()):
        items = groups[main_word]
        for i, item in enumerate(items):
            results.append([idx, main_word if i == 0 else '', item['副词'] if i == 0 else '', item['次副词'] if i == 0 else '', item['搜索词'], item['搜索量']])
            idx += 1
    return results

workbook = xlsxwriter.Workbook(output_path)
header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': '#FFFFFF', 'align': 'center', 'valign': 'vcenter'})

# 1. 源文件
ws_source = workbook.add_worksheet('源文件')
ws_source.write_row(0, 0, headers, header_fmt)
for i, row in enumerate(rows, 1):
    ws_source.write_row(i, 0, row)

# 2. 筛选（按周搜索量降序）
ws_filter = workbook.add_worksheet('筛选')
ws_filter.write_row(0, 0, ['序号'] + headers, header_fmt)
indexed_rows = [(i+1, r, int(r[10]) if len(r) > 10 and r[10] is not None else 0) for i, r in enumerate(rows)]
indexed_rows.sort(key=lambda x: -x[2])
for i, (idx, row, vol) in enumerate(indexed_rows, 1):
    ws_filter.write_row(i, 0, [idx] + row)

# 3. 筛选后词
all_words = defaultdict(lambda: {'count': 0, 'volume': 0})
for kw, vol in zip(keywords, volumes):
    words = re.findall(r"[a-z']+", kw.lower())
    for w in words:
        if len(w) > 1:
            all_words[w]['count'] += 1
            all_words[w]['volume'] += vol

ws_words = workbook.add_worksheet('筛选后词')
ws_words.write_row(0, 0, ['词', '出现次数', '搜索量'], header_fmt)
for i, (w, stats) in enumerate(sorted(all_words.items(), key=lambda x: -x[1]['count']), 1):
    ws_words.write_row(i, 0, [w, stats['count'], stats['volume']])

# 4-12. 各维度
dim_stats = {}
for dim_name, dim_config in dimensions.items():
    data = extract_dimension_data(kw_data, dim_config)
    dim_stats[dim_name] = len(data)
    if data:
        sheet_name = dim_name[:31]
        ws_dim = workbook.add_worksheet(sheet_name)
        ws_dim.write_row(0, 0, ['序号', '主词', '副词', '次副词', '搜索词', '搜索量'], header_fmt)
        for i, row in enumerate(data, 1):
            ws_dim.write_row(i, 0, row)

workbook.close()

print('桌布关键词拆词 V2 完成！')
print(f'文件保存至: {output_path}')
print(f'总关键词数: {len(keywords)}')
print(f'总周搜索量: {sum(volumes):,}')
print('各维度统计:')
for dim_name, count in dim_stats.items():
    print(f'  {dim_name}: {count} 条')

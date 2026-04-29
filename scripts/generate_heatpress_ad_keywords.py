import os, json
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 热压机 L级别投流词分层脚本
# ============================================================

INPUT_EXCEL = '/root/.openclaw/workspace/downloads/19dd6e15-23b2-8b48-8000-00009c143a54_variantExtendKeyword_zygpclelyvg1777426562140_1777426568876.xlsx'
OUTPUT_DIR = '/root/.openclaw/workspace/keyword_library/_广告词库/L级别投流词'
OUTPUT_FILE = f'{OUTPUT_DIR}/热压机_HeatPress_L级别投流词分层.xlsx'

os.makedirs(OUTPUT_DIR, exist_ok=True)

# 读取竞品拓词
from openpyxl import load_workbook
wb = load_workbook(INPUT_EXCEL)
ws = wb.active

# 表头在第二行
headers = [cell.value for cell in ws[2]]
print(f"Headers: {headers}")

data = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    d = dict(zip(headers, row))
    # 解析数值
    try:
        sv = d.get('周搜索量', 0)
        if sv is None or sv == '':
            sv = 0
        else:
            sv = int(float(sv))
    except:
        sv = 0
    
    try:
        rank = d.get('排名', None)
        if rank is None or rank == '':
            rank = 9999999
        else:
            rank = int(float(rank))
    except:
        rank = 9999999
    
    d['_sv'] = sv
    d['_rank'] = rank
    data.append(d)

print(f"Total keywords: {len(data)}")

# ============================================================
# L级别分层
# ============================================================

def assign_level(rank, sv):
    """
    标准：
    L1: rank <= 1000 或 sv >= 5000
    L2: rank 1001-30000 或 sv 500-4999
    L3: rank 30001-150000 或 sv 100-499
    L4: rank 150001-500000 或 sv 10-99
    L5: rank > 500000 或 sv < 10
    """
    if rank <= 1000 or sv >= 5000:
        return 'L1', '锚定层'
    elif rank <= 30000 or sv >= 500:
        return 'L2', '主战场'
    elif rank <= 150000 or sv >= 100:
        return 'L3', '验证区'
    elif rank <= 500000 or sv >= 10:
        return 'L4', '长尾区'
    else:
        return 'L5', '探索区'

for d in data:
    level, name = assign_level(d['_rank'], d['_sv'])
    d['_level'] = level
    d['_level_name'] = name

# 统计
level_stats = {}
for d in data:
    l = d['_level']
    if l not in level_stats:
        level_stats[l] = {'count': 0, 'sv': 0, 'name': d['_level_name']}
    level_stats[l]['count'] += 1
    level_stats[l]['sv'] += d['_sv']

print("\n分层统计:")
for l in ['L1', 'L2', 'L3', 'L4', 'L5']:
    s = level_stats.get(l, {'count': 0, 'sv': 0, 'name': ''})
    print(f"  {l} {s['name']}: {s['count']} 词, 总搜索量 {s['sv']:,}")

# ============================================================
# 生成投放词库 Excel
# ============================================================

wb_out = Workbook()
wb_out.remove(wb_out.active)

# --- Sheet 0: 分层概览 ---
ws_overview = wb_out.create_sheet("0-分层概览")
overview_headers = ['ABA分层', '排名区间', '关键词数量', '总搜索量', '流量属性', '投放策略', '预算占比建议', '匹配方式', '竞价建议']
ws_overview.append(overview_headers)

overview_rows = [
    ('L1锚定层', '1-1,000', level_stats.get('L1', {}).get('count', 0), level_stats.get('L1', {}).get('sv', 0),
     '高流量核心入口', '精准匹配+高竞价抢Top3', '40%', '精准匹配', '高竞价'),
    ('L2主战场', '1,001-30,000', level_stats.get('L2', {}).get('count', 0), level_stats.get('L2', {}).get('sv', 0),
     '高流量主战场', '词组匹配，分时段加价放量', '35%', '词组匹配', '中竞价'),
    ('L3验证区', '30,001-150,000', level_stats.get('L3', {}).get('count', 0), level_stats.get('L3', {}).get('sv', 0),
     '中流量验证区', '广泛匹配低竞价，跑转化养数据', '20%', '广泛匹配', '低竞价'),
    ('L4长尾区', '150,001-500,000', level_stats.get('L4', {}).get('count', 0), level_stats.get('L4', {}).get('sv', 0),
     '低流量长尾', '广泛匹配捡漏，找稳定出单词', '4%', '广泛匹配', '低竞价'),
    ('L5探索区', '500,001+', level_stats.get('L5', {}).get('count', 0), level_stats.get('L5', {}).get('sv', 0),
     '极低流量探索', '自动广告跑词，定期否定/打捞', '1%', '自动广告', '极低竞价'),
]

for row in overview_rows:
    ws_overview.append(list(row))

# 样式
for cell in ws_overview[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# 各层级投放策略映射
level_strategy = {
    'L1': ('精准匹配+高竞价抢Top3', '精准匹配', '高竞价'),
    'L2': ('词组匹配，分时段加价放量', '词组匹配', '中竞价'),
    'L3': ('广泛匹配低竞价，跑转化养数据', '广泛匹配', '低竞价'),
    'L4': ('广泛匹配捡漏，找稳定出单词', '广泛匹配', '低竞价'),
    'L5': ('自动广告跑词，定期否定/打捞', '自动广告', '极低竞价'),
}

# --- 各层级 Sheet ---
for l in ['L1', 'L2', 'L3', 'L4', 'L5']:
    level_data = [d for d in data if d['_level'] == l]
    level_data.sort(key=lambda x: x['_sv'], reverse=True)
    
    sheet_name = f"{l}-{level_stats.get(l, {}).get('name', l)}"
    ws = wb_out.create_sheet(sheet_name)
    
    ws.append(['关键词', '翻译', '搜索量', '排名', '相关性', '投放状态', '投放策略建议'])
    
    strategy, match_type, bid = level_strategy[l]
    
    for d in level_data:
        ws.append([
            d.get('关键词', ''),
            d.get('翻译', ''),
            d['_sv'],
            d['_rank'] if d['_rank'] < 9999999 else 'N/A',
            d.get('相关性-自动', ''),
            '待投放',
            strategy
        ])
    
    # 样式
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 列宽
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

# 概览Sheet列宽
for col in ws_overview.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        except:
            pass
    ws_overview.column_dimensions[col_letter].width = min(max_len + 2, 40)

wb_out.save(OUTPUT_FILE)
print(f"\n✅ 投放词库已保存: {OUTPUT_FILE}")

# ============================================================
# 同时更新 投流词分层结果汇总.md
# ============================================================
summary_md = f"""# L级别投流词分层结果汇总

## 热压机 HeatPress

| 层级 | 排名区间 | 关键词数 | 总搜索量 | 流量属性 | 投放策略 | 预算占比 | 匹配方式 |
|------|---------|---------|---------|---------|---------|---------|---------|
| L1 锚定层 | 1-1,000 | {level_stats.get('L1', {}).get('count', 0)} | {level_stats.get('L1', {}).get('sv', 0):,} | 高流量核心入口 | 精准匹配+高竞价抢Top3 | 40% | 精准匹配 |
| L2 主战场 | 1,001-30,000 | {level_stats.get('L2', {}).get('count', 0)} | {level_stats.get('L2', {}).get('sv', 0):,} | 高流量主战场 | 词组匹配，分时段加价放量 | 35% | 词组匹配 |
| L3 验证区 | 30,001-150,000 | {level_stats.get('L3', {}).get('count', 0)} | {level_stats.get('L3', {}).get('sv', 0):,} | 中流量验证区 | 广泛匹配低竞价，跑转化养数据 | 20% | 广泛匹配 |
| L4 长尾区 | 150,001-500,000 | {level_stats.get('L4', {}).get('count', 0)} | {level_stats.get('L4', {}).get('sv', 0):,} | 低流量长尾 | 广泛匹配捡漏，找稳定出单词 | 4% | 广泛匹配 |
| L5 探索区 | 500,001+ | {level_stats.get('L5', {}).get('count', 0)} | {level_stats.get('L5', {}).get('sv', 0):,} | 极低流量探索 | 自动广告跑词，定期否定/打捞 | 1% | 自动广告 |

**总词数**: {len(data)} | **总搜索量**: {sum(d['_sv'] for d in data):,}

### L1 Top 10 投放词
"""

l1_top = [d for d in data if d['_level'] == 'L1']
l1_top.sort(key=lambda x: x['_sv'], reverse=True)
for d in l1_top[:10]:
    summary_md += f"- **{d.get('关键词', '')}** | SV: {d['_sv']:,} | Rank: {d['_rank']:,} | {d.get('翻译', '')}\n"

summary_md += f"""
### L2 Top 10 投放词
"""
l2_top = [d for d in data if d['_level'] == 'L2']
l2_top.sort(key=lambda x: x['_sv'], reverse=True)
for d in l2_top[:10]:
    summary_md += f"- **{d.get('关键词', '')}** | SV: {d['_sv']:,} | Rank: {d['_rank']:,} | {d.get('翻译', '')}\n"

# 追加到汇总MD
summary_path = '/root/.openclaw/workspace/keyword_library/_广告词库/L级别投流词/投流词分层结果汇总.md'
with open(summary_path, 'a', encoding='utf-8') as f:
    f.write('\n\n---\n\n')
    f.write(summary_md)

print(f"✅ 汇总MD已追加: {summary_path}")
print("\n=== 全部完成 ===")

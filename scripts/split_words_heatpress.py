import json, re, os, math
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 热压机 Heat Press 关键词拆词脚本
# ============================================================

INPUT_COMPETITOR = '/root/.openclaw/workspace/downloads/19dd6e15-23b2-8b48-8000-00009c143a54_variantExtendKeyword_zygpclelyvg1777426562140_1777426568876.xlsx'
INPUT_BSR = '/root/.openclaw/workspace/downloads/19dd6e1b-22c2-8b8a-8000-0000448b147a_BSR_Heat-Press-Machines_Current_-3-US-20260429.xlsx'
OUTPUT_BASE = '/root/.openclaw/workspace/keyword_library/热压机_HeatPress'

os.makedirs(f"{OUTPUT_BASE}/01_原始数据", exist_ok=True)
os.makedirs(f"{OUTPUT_BASE}/02_拆词结果", exist_ok=True)
os.makedirs(f"{OUTPUT_BASE}/03_结构化词库", exist_ok=True)
os.makedirs(f"{OUTPUT_BASE}/04_埋词素材", exist_ok=True)

# ============================================================
# 1. 读取数据
# ============================================================
print("=== 1. 读取竞品拓词 ===")
wb_comp = load_workbook(INPUT_COMPETITOR)
ws_comp = wb_comp.active
comp_data = []
headers = [cell.value for cell in ws_comp[2]]  # 第二行是表头
for row in ws_comp.iter_rows(min_row=3, max_row=ws_comp.max_row, values_only=True):
    comp_data.append(dict(zip(headers, row)))

print(f"竞品拓词: {len(comp_data)} 条")

# 读取BSR
print("=== 2. 读取BSR ===")
wb_bsr = load_workbook(INPUT_BSR)
ws_bsr = wb_bsr.active
bsr_cols = [cell.value for cell in ws_bsr[1]]
title_idx = bsr_cols.index('商品标题')
selling_idx = bsr_cols.index('产品卖点')
brand_idx = bsr_cols.index('品牌')
asin_idx = bsr_cols.index('ASIN')
rank_idx = bsr_cols.index('小类BSR')

bsr_data = []
for row in ws_bsr.iter_rows(min_row=2, max_row=ws_bsr.max_row, values_only=True):
    bsr_data.append({
        'ASIN': row[asin_idx],
        '品牌': row[brand_idx],
        'BSR': row[rank_idx],
        '标题': row[title_idx] or '',
        '卖点': row[selling_idx] or ''
    })

print(f"BSR产品: {len(bsr_data)} 个")
for b in bsr_data:
    print(f"  #{b['BSR']} {b['品牌']} ({b['ASIN']})")

# ============================================================
# 2. BSR高频词提取
# ============================================================
print("=== 3. BSR高频词提取 ===")

def tokenize(text):
    """提取有意义的英文单词"""
    text = text.lower()
    # 保留数字和字母的组合
    tokens = re.findall(r"[a-z]+(?:[-'][a-z]+)*|\d+(?:[-.]\d+)*", text)
    # 过滤停用词
    stopwords = {
        'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by',
        'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did',
        'will', 'would', 'could', 'should', 'may', 'might', 'must', 'can', 'this', 'that', 'these',
        'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they', 'me', 'him', 'her', 'us', 'them',
        'my', 'your', 'his', 'its', 'our', 'their', 'from', 'up', 'about', 'into', 'through',
        'during', 'before', 'after', 'above', 'below', 'between', 'among', 'within', 'without',
        'as', 'if', 'so', 'than', 'too', 'very', 'just', 'now', 'then', 'here', 'there', 'when',
        'where', 'why', 'how', 'all', 'each', 'every', 'both', 'few', 'more', 'most', 'other',
        'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very',
        'also', 'etc', 'e.g', 'ie', 'per', 'approx', 'about', 'approximate', 'please', 'note',
        'tips', 'warning', 'notice', 'warm', 'reminder', 'caution', 'careful', ' carefully',
        'buy', 'purchase', 'get', 'want', 'need', 'use', 'used', 'using', 'make', 'made',
        'come', 'comes', 'coming', 'came', 'take', 'takes', 'taking', 'took', 'give', 'gives',
        'giving', 'gave', 'see', 'seen', 'saw', 'know', 'knew', 'known', 'knowns', 'think',
        'thought', 'thoughts', 'thoughtful', 'thoughtless', 'thoughtlessly', 'thoughtlessness',
        'say', 'said', 'says', 'saying', 'go', 'goes', 'going', 'went', 'gone', 'let', 'lets',
        'letting', 'put', 'puts', 'putting', 'find', 'finds', 'finding', 'found', 'tell', 'tells',
        'telling', 'told', 'ask', 'asks', 'asking', 'asked', 'seem', 'seems', 'seeming', 'seemed',
        'feel', 'feels', 'feeling', 'felt', 'try', 'tries', 'trying', 'tried', 'leave', 'leaves',
        'leaving', 'left', 'call', 'calls', 'calling', 'called', 'keep', 'keeps', 'keeping', 'kept',
        'bring', 'brings', 'bringing', 'brought', 'begin', 'begins', 'beginning', 'began', 'begun',
        'show', 'shows', 'showing', 'showed', 'shown', 'hear', 'hears', 'hearing', 'heard',
        'play', 'plays', 'playing', 'played', 'run', 'runs', 'running', 'ran', 'move', 'moves',
        'moving', 'moved', 'live', 'lives', 'living', 'lived', 'believe', 'believes', 'believing',
        'believed', 'hold', 'holds', 'holding', 'held', 'bring', 'brings', 'bringing', 'brought',
        'happen', 'happens', 'happening', 'happened', 'write', 'writes', 'writing', 'wrote', 'written',
        'sit', 'sits', 'sitting', 'sat', 'stand', 'stands', 'standing', 'stood', 'lose', 'loses',
        'losing', 'lost', 'pay', 'pays', 'paying', 'paid', 'meet', 'meets', 'meeting', 'met',
        'include', 'includes', 'including', 'included', 'continue', 'continues', 'continuing', 'continued',
        'set', 'sets', 'setting', 'follow', 'follows', 'following', 'followed', 'stop', 'stops',
        'stopping', 'stopped', 'create', 'creates', 'creating', 'created', 'speak', 'speaks',
        'speaking', 'spoke', 'spoken', 'read', 'reads', 'reading', 'allow', 'allows', 'allowing',
        'allowed', 'add', 'adds', 'adding', 'added', 'spend', 'spends', 'spending', 'spent',
        'grow', 'grows', 'growing', 'grew', 'grown', 'open', 'opens', 'opening', 'opened',
        'walk', 'walks', 'walking', 'walked', 'win', 'wins', 'winning', 'won', 'offer', 'offers',
        'offering', 'offered', 'remember', 'remembers', 'remembering', 'remembered', 'love',
        'loves', 'loving', 'loved', 'consider', 'considers', 'considering', 'considered', 'appear',
        'appears', 'appearing', 'appeared', 'buy', 'buys', 'buying', 'bought', 'wait', 'waits',
        'waiting', 'waited', 'serve', 'serves', 'serving', 'served', 'die', 'dies', 'dying', 'died',
        'send', 'sends', 'sending', 'sent', 'expect', 'expects', 'expecting', 'expected', 'build',
        'builds', 'building', 'built', 'stay', 'stays', 'staying', 'stayed', 'fall', 'falls',
        'falling', 'fell', 'fallen', 'cut', 'cuts', 'cutting', 'reach', 'reaches', 'reaching',
        'reached', 'kill', 'kills', 'killing', 'killed', 'remain', 'remains', 'remaining', 'remained',
        'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 'ten', 'first',
        'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh', 'eighth', 'ninth', 'tenth',
        'x', 'y', 'z', 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm',
        'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z',
    }
    filtered = []
    for t in tokens:
        if len(t) < 2 and not t.isdigit():
            continue
        if t in stopwords:
            continue
        filtered.append(t)
    return filtered

# BSR标题+卖点分词
bsr_all_text = ' '.join([b['标题'] + ' ' + b['卖点'] for b in bsr_data])
bsr_tokens = tokenize(bsr_all_text)
bsr_counter = Counter(bsr_tokens)
bsr_top_words = bsr_counter.most_common(100)

print("BSR Top 20高频词:")
for word, cnt in bsr_top_words[:20]:
    print(f"  {word}: {cnt}")

# ============================================================
# 3. VAN五层分层
# ============================================================
print("=== 4. VAN五层分层 ===")

def get_search_volume(row):
    try:
        sv = row.get('周搜索量', 0)
        if sv is None:
            return 0
        return int(float(sv))
    except:
        return 0

def get_rank(row):
    try:
        r = row.get('排名', None)
        if r is None or r == '':
            return 9999999
        return int(float(r))
    except:
        return 9999999

# 给每个词计算分层
for row in comp_data:
    row['_sv'] = get_search_volume(row)
    row['_rank'] = get_rank(row)
    sv = row['_sv']
    rank = row['_rank']
    
    if rank <= 1000 or sv >= 5000:
        row['_level'] = 'L1'
        row['_level_name'] = '锚定层'
    elif rank <= 30000 or sv >= 500:
        row['_level'] = 'L2'
        row['_level_name'] = '主战场'
    elif rank <= 150000 or sv >= 100:
        row['_level'] = 'L3'
        row['_level_name'] = '验证区'
    elif rank <= 500000 or sv >= 10:
        row['_level'] = 'L4'
        row['_level_name'] = '长尾区'
    else:
        row['_level'] = 'L5'
        row['_level_name'] = '探索区'

level_stats = {}
for row in comp_data:
    l = row['_level']
    level_stats[l] = level_stats.get(l, {'count': 0, 'sv': 0})
    level_stats[l]['count'] += 1
    level_stats[l]['sv'] += row['_sv']

for l in ['L1', 'L2', 'L3', 'L4', 'L5']:
    s = level_stats.get(l, {'count': 0, 'sv': 0})
    name = {'L1':'锚定层', 'L2':'主战场', 'L3':'验证区', 'L4':'长尾区', 'L5':'探索区'}[l]
    print(f"  {l} {name}: {s['count']} 词, 总搜索量 {s['sv']:,}")

# ============================================================
# 4. 15维度拆词
# ============================================================
print("=== 5. 15维度拆词 ===")

# 热压机类目维度词典
dimensions = {
    '尺寸规格': {
        'words': [
            'mini', 'small', 'portable', 'tiny', 'compact', 'large', 'big', '12x10', '15x15', 
            '9x9', '10x10', '12x12', '15x15', '16x20', '16x24', '20x20', '38x38', '40x50',
            '12 inch', '15 inch', '16 inch', '20 inch', '38 cm', '40 cm', '50 cm',
            'inch', 'cm', 'mm', 'feet', 'foot'
        ],
        'core_terms': ['heat press', 'heat press machine', 'press', 'machine', 'iron']
    },
    '产品类型': {
        'words': [
            'heat press', 'heat press machine', 'heat presser', 'heat transfer', 'heat transfer machine',
            't shirt press', 'shirt press', 't-shirt press', 'tshirt press',
            'iron press', 'iron press machine', 'mini iron', 'craft iron', 'flat iron',
            'sublimation machine', 'sublimation press', 'vinyl press', 'vinyl cutter',
            'printing press', 'screen printing', 'dtg printer', 'direct to garment',
            'mug press', 'hat press', 'cap press', 'plate press', 'combo press',
            'clamshell', 'swing away', 'drawer', 'slide out'
        ],
        'core_terms': ['heat press', 'heat press machine']
    },
    '功能特性': {
        'words': [
            'sublimation', 'vinyl', 'htv', 'heat transfer vinyl', 'transfer paper',
            'infusible ink', 'transfer tape', 'adhesive vinyl', 'permanent vinyl',
            'temperature control', 'temp control', 'adjustable temperature', 'adjustable temp',
            'digital display', 'led display', 'lcd display', 'timer', 'auto shutoff',
            'auto shut off', 'auto power off', 'overheat protection', 'pressure adjustment',
            'pressure adjustable', 'dual heating', 'dual plate', 'dual function',
            'quick heat', 'fast heating', 'rapid heat', 'even heat', 'uniform heat',
            'teflon coated', 'non stick', 'non-stick', 'detachable', 'removable',
            'rotating', 'swiveling', '360 degree', '360°', 'reversible'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    },
    '温度设置': {
        'words': [
            'low temp', 'medium temp', 'high temp', 'low temperature', 'medium temperature',
            'high temperature', '3 temp', '3 temperature', '3 heat', '3 settings',
            'adjustable temp', 'adjustable temperature', 'variable temp', 'variable temperature',
            'preset temp', 'preset temperature', '284', '320', '356', '140', '160', '180',
            '170', '180', '200', '210', '230', '250', '300', '350', '400',
            'fahrenheit', 'celsius', '°f', '°c'
        ],
        'core_terms': ['heat press', 'heat press machine']
    },
    '使用场景': {
        'words': [
            't shirts', 't-shirts', 'tshirts', 'shirt', 'clothing', 'apparel', 'fabric',
            'hat', 'hats', 'cap', 'caps', 'beanie', 'mug', 'mugs', 'cup', 'cups',
            'plate', 'plates', 'bag', 'bags', 'tote bag', 'pillow', 'pillows', 'case',
            'phone case', 'mask', 'masks', 'sock', 'socks', 'shoe', 'shoes',
            'wood', 'metal', 'glass', 'ceramic', 'canvas', 'leather', 'polyester',
            'cotton', 'nylon', 'spandex', 'linen', 'wool', 'silk', 'plush', 'velvet',
            'home', 'home use', 'business', 'commercial', 'professional', 'industrial',
            'shop', 'store', 'studio', 'workshop', 'garage', 'office', 'craft room'
        ],
        'core_terms': ['heat press', 'press', 'machine', 'iron']
    },
    '目标人群': {
        'words': [
            'beginner', 'beginners', 'starter', 'newbie', 'novice', 'starter kit',
            'hobbyist', 'hobby', 'crafter', 'crafters', 'craft enthusiast', 'diy',
            'do it yourself', 'home user', 'small business', 'small business owner',
            'entrepreneur', 'side hustle', 'etsy seller', 'print on demand', 'pod',
            'kids', 'children', 'adult', 'adults', 'family', 'women', 'men'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    },
    '品牌': {
        'words': [
            'htvront', 'cricut', 'siser', 'oracal', 'teckwrap', 'sublimation', 'vevor',
            'f2c', 'fancierstudio', 'powerpress', 'superdeal', 'furgle', 'bettersub',
            'honoson', 'calogy', 'shzond', 'mophorn', 'yescom', ' PromoHeat', 'gecko',
            'us cutter', 'uscutter', 'silhouette', 'brother', 'epson', 'sawgrass'
        ],
        'core_terms': ['heat press', 'press', 'machine', 'vinyl', 'sublimation']
    },
    '颜色': {
        'words': [
            'white', 'black', 'pink', 'blue', 'green', 'red', 'yellow', 'purple', 'orange',
            'mint', 'mint green', 'rose', 'rose gold', 'gold', 'silver', 'gray', 'grey',
            'teal', 'navy', 'turquoise', 'lavender', 'peach', 'coral', 'maroon', 'burgundy',
            'charcoal', 'beige', 'cream', 'ivory'
        ],
        'core_terms': ['heat press', 'press', 'machine', 'iron']
    },
    '配件套餐': {
        'words': [
            'bundle', 'kit', 'starter kit', 'combo', 'set', 'package', 'with accessories',
            'with tools', 'with paper', 'with vinyl', 'with tape', 'with teflon',
            'with sheets', 'with mat', 'with pad', 'with platen', 'extra plate',
            'additional plate', 'replacement', 'spare parts'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    },
    '价格定位': {
        'words': [
            'cheap', 'affordable', 'budget', 'inexpensive', 'low cost', 'economical',
            'expensive', 'premium', 'high end', 'high-end', 'luxury', 'professional grade',
            'value', 'deal', 'sale', 'discount', 'under', 'below', 'cheap price',
            'best price', 'low price', 'best value', 'cost effective'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    },
    '电源电压': {
        'words': [
            '110v', '220v', '110 volt', '220 volt', '110 v', '220 v', 'dual voltage',
            'us plug', 'eu plug', 'uk plug', 'au plug', 'us standard', 'european',
            'cord', 'cable', 'wire', 'power cord', 'power supply', 'ac', 'dc'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    },
    '操作方式': {
        'words': [
            'easy', 'simple', 'user friendly', 'user-friendly', 'one button', 'one touch',
            'automatic', 'auto', 'manual', 'handheld', 'hand held', 'portable', 'lightweight',
            'light weight', 'compact', 'ergonomic', 'comfortable', 'convenient',
            'quick', 'fast', 'instant', 'ready', 'preheat', 'pre heated', 'pre-heated'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    },
    '材质工艺': {
        'words': [
            'teflon', 'ceramic', 'aluminum', 'aluminium', 'steel', 'stainless steel',
            'silicone', 'rubber', 'foam', 'cotton', 'fabric', 'coated', 'plated',
            'durable', 'heavy duty', 'sturdy', 'solid', 'quality', 'high quality',
            'premium material', 'industrial grade', 'commercial grade'
        ],
        'core_terms': ['heat press', 'press', 'plate', 'machine']
    },
    '工艺技术': {
        'words': [
            'sublimation printing', 'dye sublimation', 'heat transfer printing',
            'vinyl cutting', 'vinyl weeding', 'screen printing', 'dtg', 'direct to garment',
            'laser transfer', 'inkjet transfer', 'pigment ink', 'dye ink',
            'infusible', 'infusible ink', 'easyweed', 'permanent vinyl', 'removable vinyl',
            'glitter vinyl', 'holographic', 'metallic', 'foil', 'flock', 'reflective'
        ],
        'core_terms': ['heat press', 'vinyl', 'sublimation', 'transfer']
    },
    '竞争属性': {
        'words': [
            'top', 'best', 'best seller', 'bestseller', 'best selling', 'popular',
            'recommended', 'highly rated', '5 star', 'review', 'reviews', 'rated',
            'amazon choice', 'amazon\'s choice', 'editor choice', 'top rated',
            'new', 'latest', '2024', '2025', '2026', 'upgraded', 'improved', 'enhanced',
            'version', 'model', 'generation', 'gen', 'pro', 'plus', 'max', 'ultra'
        ],
        'core_terms': ['heat press', 'press', 'machine']
    }
}

# 拆词匹配
def match_dimension(keyword, dim_words):
    """判断关键词是否匹配某维度"""
    kw_lower = keyword.lower()
    for word in dim_words:
        if word in kw_lower:
            return True
    return False

dim_results = {k: [] for k in dimensions.keys()}

for row in comp_data:
    kw = row.get('关键词', '') or ''
    matched_any = False
    for dim_name, dim_cfg in dimensions.items():
        if match_dimension(kw, dim_cfg['words']):
            dim_results[dim_name].append(row)
            matched_any = True
    # 未匹配任何维度的归入"其他"
    if not matched_any:
        pass  # 不收集

print("\n各维度覆盖情况:")
dim_summary = {}
for dim_name, rows in dim_results.items():
    total_sv = sum(r['_sv'] for r in rows)
    dim_summary[dim_name] = {'count': len(rows), 'sv': total_sv}
    print(f"  {dim_name}: {len(rows)} 词, 总搜索量 {total_sv:,}")

# ============================================================
# 5. 输出Excel拆词结果
# ============================================================
print("=== 6. 输出Excel ===")

wb_out = Workbook()
wb_out.remove(wb_out.active)  # 删除默认sheet

# Sheet 1: 源文件（按搜索量排序）
ws_src = wb_out.create_sheet("源文件")
headers = ['序号', '关键词', '翻译', '相关性', '周搜索量', '排名', 'L层级', '层级名称']
ws_src.append(headers)
for i, row in enumerate(sorted(comp_data, key=lambda x: x['_sv'], reverse=True), 1):
    ws_src.append([
        i,
        row.get('关键词', ''),
        row.get('翻译', ''),
        row.get('相关性-自动', ''),
        row['_sv'],
        row['_rank'],
        row['_level'],
        row['_level_name']
    ])

# Sheet 2: VAN分层汇总
ws_level = wb_out.create_sheet("VAN分层")
ws_level.append(['层级', '层级名称', '关键词数', '总搜索量', '代表词'])
for l in ['L1', 'L2', 'L3', 'L4', 'L5']:
    s = level_stats.get(l, {'count': 0, 'sv': 0})
    name = {'L1':'锚定层', 'L2':'主战场', 'L3':'验证区', 'L4':'长尾区', 'L5':'探索区'}[l]
    # 找该层级搜索量top3
    top_words = sorted([r for r in comp_data if r['_level']==l], key=lambda x: x['_sv'], reverse=True)[:3]
    rep = ', '.join([r.get('关键词','') for r in top_words])
    ws_level.append([l, name, s['count'], s['sv'], rep])

# Sheet 3: 筛选（搜索量>0）
ws_filter = wb_out.create_sheet("筛选")
ws_filter.append(['序号', '关键词', '翻译', '周搜索量', '排名', 'L层级'])
filtered = [r for r in comp_data if r['_sv'] > 0]
for i, row in enumerate(sorted(filtered, key=lambda x: x['_sv'], reverse=True), 1):
    ws_filter.append([i, row.get('关键词',''), row.get('翻译',''), row['_sv'], row['_rank'], row['_level']])

# Sheet 4+: 各维度拆词
for dim_name in dimensions.keys():
    ws_dim = wb_out.create_sheet(dim_name[:20])  # sheet名限制31字符
    ws_dim.append(['序号', '主词', '副词', '次副词', '搜索词', '搜索量', 'L层级', '翻译'])
    rows = sorted(dim_results[dim_name], key=lambda x: x['_sv'], reverse=True)
    
    # 找出每个词的"主词"（匹配到的维度词中优先级最高的）
    for i, row in enumerate(rows, 1):
        kw = row.get('关键词', '')
        # 找匹配到的维度词
        matched_words = [w for w in dimensions[dim_name]['words'] if w in kw.lower()]
        main = matched_words[0] if matched_words else ''
        # 其余作为副词
        sub = ', '.join(matched_words[1:3]) if len(matched_words) > 1 else ''
        ws_dim.append([
            i, main, sub, '', kw, row['_sv'], row['_level'], row.get('翻译','')
        ])

# BSR高频词 Sheet
ws_bsr_words = wb_out.create_sheet("BSR高频词")
ws_bsr_words.append(['序号', '词', '出现次数', '来源'])
for i, (word, cnt) in enumerate(bsr_top_words[:50], 1):
    ws_bsr_words.append([i, word, cnt, 'BSR Top3标题+卖点'])

# 样式设置
def set_header_style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

for ws in wb_out.worksheets:
    set_header_style(ws)
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

output_excel = f"{OUTPUT_BASE}/02_拆词结果/热压机关键词拆词_完整层级V1.xlsx"
wb_out.save(output_excel)
print(f"Excel已保存: {output_excel}")

# ============================================================
# 6. 输出结构化JSON词库
# ============================================================
print("=== 7. 输出JSON词库 ===")

json_lib = {
    "category": "热压机_HeatPress",
    "created": "2026-04-29",
    "source": {
        "competitor_asins": 50,
        "competitor_keywords": len(comp_data),
        "bsr_products": len(bsr_data)
    },
    "van_levels": {},
    "dimensions": {},
    "bsr_top_words": [],
    "title_formula": {
        "structure": "[Brand] + [Size/Type] + [Core Product] + [Key Feature] + [Target/Scene] + [Color]",
        "example": "Mini Heat Press Machine for T-Shirts, Portable Small Iron Press with 3 Temp Settings, Easy to Use for DIY Crafts and Heat Transfer Vinyl Projects"
    },
    "bullet_strategy": [
        "尺寸规格 + 温度设置",
        "功能特性 + 适用材质",
        "操作便捷 + 安全保护",
        "目标人群 + 使用场景",
        "配件内容 + 售后保障"
    ],
    "backend_search_terms": "sublimation vinyl, heat transfer paper, tshirt press machine, craft iron mini, portable heat press small"
}

for l in ['L1', 'L2', 'L3', 'L4', 'L5']:
    s = level_stats.get(l, {'count': 0, 'sv': 0})
    name = {'L1':'锚定层', 'L2':'主战场', 'L3':'验证区', 'L4':'长尾区', 'L5':'探索区'}[l]
    top_words = sorted([r for r in comp_data if r['_level']==l], key=lambda x: x['_sv'], reverse=True)[:10]
    json_lib['van_levels'][l] = {
        'name': name,
        'count': s['count'],
        'total_search_volume': s['sv'],
        'top_keywords': [
            {'keyword': r.get('关键词',''), 'sv': r['_sv'], 'rank': r['_rank']}
            for r in top_words
        ]
    }

for dim_name in dimensions.keys():
    rows = dim_results[dim_name]
    total_sv = sum(r['_sv'] for r in rows)
    # 维度内高频词统计
    dim_word_counter = Counter()
    for r in rows:
        kw = r.get('关键词', '')
        for w in dimensions[dim_name]['words']:
            if w in kw.lower():
                dim_word_counter[w] += r['_sv']  # 按搜索量加权
    
    json_lib['dimensions'][dim_name] = {
        'count': len(rows),
        'total_search_volume': total_sv,
        'top_attributes': [
            {'word': w, 'weighted_sv': cnt}
            for w, cnt in dim_word_counter.most_common(10)
        ],
        'keywords': [
            {'keyword': r.get('关键词',''), 'sv': r['_sv'], 'rank': r['_rank']}
            for r in sorted(rows, key=lambda x: x['_sv'], reverse=True)[:20]
        ]
    }

json_lib['bsr_top_words'] = [
    {'word': w, 'count': c} for w, c in bsr_top_words[:30]
]

json_path = f"{OUTPUT_BASE}/03_结构化词库/热压机埋词库.json"
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(json_lib, f, ensure_ascii=False, indent=2)

print(f"JSON已保存: {json_path}")

# ============================================================
# 7. 输出MD埋词素材
# ============================================================
print("=== 8. 输出MD埋词素材 ===")

md_content = f"""# 热压机 Heat Press SEO关键词词库

**创建日期**: 2026-04-29
**数据来源**: 50个竞品ASIN拓词 + BSR Top 3产品

---

## VAN五层流量模型

| 层级 | 名称 | 关键词数 | 总搜索量 | 匹配方式 | 预算占比 |
|------|------|---------|---------|---------|---------|
| L1 | 锚定层 | {level_stats.get('L1', {}).get('count', 0)} | {level_stats.get('L1', {}).get('sv', 0):,} | 精准匹配+高竞价 | 40% |
| L2 | 主战场 | {level_stats.get('L2', {}).get('count', 0)} | {level_stats.get('L2', {}).get('sv', 0):,} | 词组匹配分时段加价 | 35% |
| L3 | 验证区 | {level_stats.get('L3', {}).get('count', 0)} | {level_stats.get('L3', {}).get('sv', 0):,} | 广泛匹配低竞价养数据 | 20% |
| L4 | 长尾区 | {level_stats.get('L4', {}).get('count', 0)} | {level_stats.get('L4', {}).get('sv', 0):,} | 广泛匹配捡漏 | 4% |
| L5 | 探索区 | {level_stats.get('L5', {}).get('count', 0)} | {level_stats.get('L5', {}).get('sv', 0):,} | 自动广告跑词 | 1% |

### L1 锚定层 Top 10
"""

l1_top = sorted([r for r in comp_data if r['_level']=='L1'], key=lambda x: x['_sv'], reverse=True)[:10]
for r in l1_top:
    md_content += f"- **{r.get('关键词','')}** | SV: {r['_sv']:,} | Rank: {r['_rank']:,} | {r.get('翻译','')}\n"

md_content += f"""
### L2 主战场 Top 10
"""
l2_top = sorted([r for r in comp_data if r['_level']=='L2'], key=lambda x: x['_sv'], reverse=True)[:10]
for r in l2_top:
    md_content += f"- **{r.get('关键词','')}** | SV: {r['_sv']:,} | Rank: {r['_rank']:,} | {r.get('翻译','')}\n"

md_content += f"""
---

## 15维度拆词覆盖

"""

for dim_name in dimensions.keys():
    s = dim_summary.get(dim_name, {'count':0, 'sv':0})
    md_content += f"""### {dim_name}
- **覆盖关键词**: {s['count']} 个
- **总搜索量**: {s['sv']:,}
- **Top属性词**:\n"""
    # 维度内Top属性词
    dim_word_counter = Counter()
    for r in dim_results[dim_name]:
        kw = r.get('关键词', '')
        for w in dimensions[dim_name]['words']:
            if w in kw.lower():
                dim_word_counter[w] += r['_sv']
    for w, cnt in dim_word_counter.most_common(5):
        md_content += f"  - {w}: {cnt:,} SV\n"
    md_content += "\n"

md_content += f"""---

## BSR Top 3 产品高频词

"""
for i, (word, cnt) in enumerate(bsr_top_words[:20], 1):
    md_content += f"{i}. **{word}** — {cnt}次\n"

md_content += f"""
---

## 埋词策略

### 标题公式
```
[Brand] + [Size/Type] + [Core Product] + [Key Feature] + [Target/Scene] + [Color]
```

**示例**: Mini Heat Press Machine for T-Shirts, Portable Small Iron Press with 3 Temp Settings, Easy to Use for DIY Crafts and Heat Transfer Vinyl Projects

### 五点埋词优先级
1. **尺寸规格 + 温度设置** — mini size, 3 heat settings, adjustable temperature
2. **功能特性 + 适用材质** — sublimation, HTV vinyl, heat transfer, cotton, polyester
3. **操作便捷 + 安全保护** — easy to use, one button, auto shutoff, safe base
4. **目标人群 + 使用场景** — beginners, crafters, DIY, t-shirts, hats, pillows
5. **配件内容 + 售后保障** — with accessories, insulated base, customer service

### 后台搜索词
- sublimation vinyl heat transfer
- tshirt press machine small
- craft iron mini portable
- heat press for beginners
- htv vinyl starter kit
- diy heat transfer projects

---

## 输出文件位置

| 子目录 | 文件 | 说明 |
|--------|------|------|
| 01_原始数据 | 竞品拓词_热压机_2026-04-29.xlsx | 原始拓词数据 |
| 01_原始数据 | BSR热压机_Top3_2026-04-29.xlsx | 原始BSR数据 |
| 02_拆词结果 | 热压机关键词拆词_完整层级V1.xlsx | **15维度完整拆词** |
| 03_结构化词库 | 热压机埋词库.json | 结构化JSON词库 |
| 04_埋词素材 | 热压机SEO关键词词库.md | 本文件 |
"""

md_path = f"{OUTPUT_BASE}/04_埋词素材/热压机SEO关键词词库.md"
with open(md_path, 'w', encoding='utf-8') as f:
    f.write(md_content)

print(f"MD已保存: {md_path}")

# 复制原始数据
import shutil
shutil.copy(INPUT_COMPETITOR, f"{OUTPUT_BASE}/01_原始数据/竞品拓词_热压机_2026-04-29.xlsx")
shutil.copy(INPUT_BSR, f"{OUTPUT_BASE}/01_原始数据/BSR热压机_Top3_2026-04-29.xlsx")

print("=== 全部完成 ===")
print(f"\n输出目录: {OUTPUT_BASE}")
print(f"  - 01_原始数据/ 竞品拓词 + BSR数据")
print(f"  - 02_拆词结果/ 热压机关键词拆词_完整层级V1.xlsx")
print(f"  - 03_结构化词库/ 热压机埋词库.json")
print(f"  - 04_埋词素材/ 热压机SEO关键词词库.md")
